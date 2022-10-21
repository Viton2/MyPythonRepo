from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Planilha de compras"
ws['A1'] = "Produto"
ws['B1'] = "Valor"
ws['C1'] = "Quantidade"
ws['D1'] = "Valor Total/Produto"
ws.column_dimensions['A'].width = 30
ws.column_dimensions['C'].width = 11
ws.column_dimensions['D'].width = 18

linha = 2
total = 0
nome = ""
while nome != 'FIM':
  nome = input("Informe o produto: [Digite FIM para terminar] ")
  if nome == 'FIM':
      continue
  valor = input("informe o valor do produto: ")
  quantidade = int(input("Informe a quantidade do produto: "))
  print("-----------------------------------------")
  total_produto = "%.2f" % float(float(valor) * quantidade)
  num = str(linha-1)
  ws.cell(row=linha, column=1, value=nome)
  ws.cell(row=linha, column=2, value=valor)
  ws.cell(row=linha, column=3, value=quantidade)
  ws.cell(row=linha, column=4, value=total_produto)
  total += (float(valor) * int(quantidade))
  linha += 1
ws.cell(row=linha, column=1, value=f'Valor total geral a ser pago: R${total :.2f}')

wb.save("c:\\temp\\carrinho.xlsx")
print ("Planilha salva em c:/temp/carrinho.xlsx")