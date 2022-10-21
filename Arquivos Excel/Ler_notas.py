from openpyxl import load_workbook

wb = load_workbook("C:\\temp\\planilha.xlsx")

ws = wb['Planilha1']

for linha in range(2, 7):
  aluno = ws.cell(column=1, row=linha).value
  nota1 = ws.cell(column=2, row=linha).value
  nota2 = ws.cell(column=3, row=linha).value
  media = ((nota1 + nota2)/2)
  print("Aluno:".rjust(11), aluno)
  print("Prova-1:".rjust(10), nota1)
  print("Prova-2:".rjust(10), nota2)
  print("Média:".rjust(8), media)
  if media < 7:
      print("Reprovado!".rjust(16))
  else:
      print("Aprovado!".rjust(16))
  print("-------------------------------------")